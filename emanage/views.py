from django.shortcuts import render,redirect,get_object_or_404
from emanage.forms import SignUpForm,LoginForm
from django.contrib.auth import login,logout,authenticate
from .models import *
from django.contrib.auth.decorators import login_required,user_passes_test
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models import Sum
from django.contrib import messages
from django.utils.datastructures import MultiValueDictKeyError
from django.core.files.storage import FileSystemStorage
import pandas as pd
import plotly.graph_objects as go
import matplotlib.pyplot as plt
import openpyxl
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.http import JsonResponse
from datetime import datetime
from itertools import groupby
from django.db.models.functions  import TruncMonth
from rest_framework import generics
from .serializers import Training_EnquirySerializer,Development_EnquirySerializer

def register(request):
    msg = None
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            msg = 'user created'
            return redirect('emanage:login_view')
        else:
            msg = 'form is not valid'
    else:
        form = SignUpForm()
    return render(request,'enquiry/register.html', {'form': form, 'msg': msg})

def index(request):
    return render(request,'enquiry/index1.html')
def login_view(request):
    form = LoginForm(request.POST or None)
    msg = None
    if request.method == 'POST':
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None and user.is_admin:
                login(request, user)
                return redirect('emanage:admin_page')
            elif user is not None and user.is_enquiry:
                login(request, user)
                return redirect('emanage:home')
            elif user is not None and user.is_account:
                login(request, user)
                return redirect('emanage:account_home')
            elif user is not None and user.is_it:
                login(request, user)
                return redirect('emanage:it_home')
            elif user is not None and user.is_hr:
                login(request, user)
                return redirect('emanage:hr_home')
            elif user is not None and user.is_training:
                login(request, user)
                return redirect('emanage:training_home')
            else:
                msg= 'invalid credentials'
        else:
            msg = 'error validating form'
    return render(request, 'enquiry/login.html', {'form': form, 'msg': msg})

def Logout(request):
    logout(request)    
    return redirect('emanage:login')

class Training_EnquiryCreateView(generics.CreateAPIView):
    queryset = Training_Enquiry.objects.all()
    serializer_class = Training_EnquirySerializer
    
class Training_EnquiryListView(generics.ListAPIView):
    queryset = Training_Enquiry.objects.all()
    serializer_class = Training_EnquirySerializer
    
class Training_EnquiryDeleteView(generics.DestroyAPIView):
    queryset = Training_Enquiry.objects.all()
    serializer_class = Training_EnquirySerializer
    lookup_field='id'

class Development_EnquiryCreateView(generics.CreateAPIView):
    queryset = Development_Enquiry.objects.all()
    serializer_class = Development_EnquirySerializer
    
class Development_EnquiryListView(generics.ListAPIView):
    queryset = Development_Enquiry.objects.all()
    serializer_class = Development_EnquirySerializer
    
class Development_EnquiryDeleteView(generics.DestroyAPIView):
    queryset = Development_Enquiry.objects.all()
    serializer_class = Development_EnquirySerializer
    lookup_field='id'
    
def training_enq_list(request):
    all_data=Training_Enquiry.objects.all()
    return render(request,'enquiry/training_list.html',context={'all_data':all_data})

def home_view(request):
    all_enquiry=Enquiry.objects.all().count()
    all_followup=reason_for_follow_up.objects.all().count()
    user=User.objects.filter(is_enquiry=request.user.is_enquiry)
    notdone= reason_for_follow_up.objects.filter(Followup_Status = 'pending').count()
    approved = reason_for_follow_up.objects.filter(Followup_Status = 'FollowUp Done').count()
    context={'all_enquiry':all_enquiry,
            'all_followup':all_followup,
            'notdone':notdone,
            'approved':approved,
            'user':user}
    return render(request,'enquiry/index.html',context)

def enquiry_list_fun(request):
    # enquiry_list=Enquiry.objects.all()
    query = request.GET.get('q', '')  # Get the search query from the request parameters

    # Retrieve all items or filter based on the search query
    if query:
        items = Enquiry.objects.filter(
            Q(Full_Name__icontains=query)  # Replace 'field1' with the actual name of the field to search
               # Replace 'field3' with another field to search, if needed
            # Add more fields to search as necessary using Q objects
        ).order_by('-id')
        
    else:
        items = Enquiry.objects.all().order_by('-id')
       
    # items=Enquiry.objects.all().order_by('-id')
    items_per_page = 9  # Number of items to display per page
    # items = Enquiry.objects.all() # Retrieve the items you want to paginate
    total_enquiries=items.count()

    paginator = Paginator(items, items_per_page)  # Create a Paginator object

    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # Get the Page object for the current page

    return render(request,'enquiry/enquiry_list.html',{'items':page_obj,'total_enquiries':total_enquiries,'page_obj':page_obj})

def is_enquiry(self):    # login decorator
    if str(self.user_type) == 'Is enquiry':
        return True
    else:
        return False
rec_login_required = user_passes_test(lambda u: True if u.is_enquiry else False, login_url='/')

def enquiry_login_required(view_func):
    decorated_view_func = login_required(rec_login_required(view_func), login_url='/')
    return decorated_view_func


   # enquiry form
def base_view(request):
        mydata = Cource.objects.all()
        # course=Cource.objects.filter(enquiry=id).last()
        if request.method == "POST":
            name = request.POST.get('Full_Name')
            gender = request.POST.get('gender')
            email = request.POST.get('Email_Address')
            mobile = request.POST.get('Contact_Number')
            address=request.POST.get('Permanent_Address')
            whatsapp = request.POST.get('Whatsapp_Number')
            qualification=request.POST.get('qualification')
            media=request.POST.get('How_did_you_hear_about_us')
            skill = request.POST.get('Technical_Skills')
            certificate = request.POST.get('Certification_Done')
            interest = request.POST.get('Area_of_Intrest')
            reference = request.POST.get('reference')
            remark = request.POST.get('remark')
            Cource_name=request.POST.get('Cource_name')
            massage=request.POST.get('massage')
            dob=request.POST.get('Date_of_Birth')
            alternate=request.POST.get('Alternate_Number')
            location=request.POST.get('Office_Location_Preference')
            mode=request.POST.get('Training_Mode')
            print("Course name",Cource_name)
            # Cname = request.POST.get('Cource_name')
            Enq_save = Enquiry(qualification=qualification,How_did_you_hear_about_us=media,Full_Name=name,gender=gender,Email_Address=email,Contact_Number=mobile,Whatsapp_Number=whatsapp,Technical_Skills=skill,Certification_Done=certificate,
                               Area_of_Intrest=interest,massage=massage,reference=reference,remark=remark,Cource_name_id=Cource_name,Alternate_Number=alternate,
                               Training_Mode=mode,Date_of_Birth=dob,Office_Location_Preference=location,Permanent_Address=address)
            mydata.Cource_name_id=Cource_name
            Enq_save.save()
            thank = True
            return redirect('/enq_list')
        else:
            return render(request, 'enquiry/enform.html',context={'mydata':mydata})
   
def enquiry_final(request): # final enquiry list
    final = Enquiry.objects.filter(enquiry_Status = 'Finalized').all()
    return render(request, 'enquiry/enq_finalize.html', context={'final':final})

def enquiry_pending(request): # pending enquiry list
    pending = Enquiry.objects.filter(enquiry_Status = 'Pending').all()
    return render(request, 'enquiry/enq_pending.html', context={'pending':pending})

def add_course(request):
    if request.method == 'POST':
        Cource_name=request.POST.get('Cource_name')
        fees=request.POST['fees']
        durations=request.POST['durations']
        course=Cource.objects.create(Cource_name=Cource_name,fees=fees,durations=durations)
        course.save()
        messages.success(request,'Course Added Successfully')
        return redirect('/courselist')
    else:
        return render(request,'enquiry/addcourse.html')
    
def course_list(request):
    all_course=Cource.objects.all()
    return render(request,'enquiry/courselist.html',{'all_course':all_course})

def updatecourse(request,id):
    course=Cource.objects.get(id=id)
    if request.method== 'POST':
        Cource_name=request.POST.get('Cource_name')
        fees=request.POST['fees']
        durations=request.POST['durations']
      
        course.Cource_name=Cource_name
        course.fees=fees
        course.durations=durations
        course.save()
        return redirect('/courselist')
    else:
        return render(request,'enquiry/updatecourse.html',context={'course':course})
    
def enquiry_profile(request):
    user = User.objects.get(id=request.user.id)
    student = Enquiry.objects.filter(user=user).first()
    context={
        "user": user,
        "student": student
    }
    return render(request, 'enquiry/profile.html', context)


def enquiry_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('enq:enquiry_profile')
    else:
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
      
        try:
            customuser =User.objects.get(id=request.user.id)
            customuser.username = username
            customuser.email = email
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()

            # student = Enquiry.objects.get(admin=customuser.id)
            # student.address = address
            # student.save()
            
            messages.success(request, "Profile Updated Successfully")
            return redirect('emanage:enquiry_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('emanage:enquiry_profile')

def follow_up_view(request,id):
    enquiry=Enquiry.objects.get(id=id)
    myreason=Reason_follow.objects.all()
    all_enquiry=Enquiry.objects.filter(id=id).only("Full_Name","Contact_Number")
    context={
        'all_enquiry':all_enquiry,'myreason':myreason,'enquiry':enquiry
    }
    if request.method == 'POST':
            if reason_for_follow_up.objects.count() >= 3:
               if reason_for_follow_up.objects.filter(enquiry_id=id).count()>=3: 
                messages.error(request, 'Sorry, you have reached the maximum number of follow-up submissions.')
                return redirect('enquiry:follwuplist',id=id)
            enquiry=enquiry
            date = request.POST['date']
            time = request.POST['time']
            remark = request.POST['remark']
            reason=request.POST['reason']
            followup=reason_for_follow_up.objects.create(enquiry=enquiry,date=date,remark=remark,time=time,reason_id=reason)
           
            myreason.reason_id=reason
            followup.save()
            messages.success(request, 'Thank you for your follow-up submission.')
            return redirect('emanage:follwuplist',id=id)
    else:
       return render(request, 'enquiry/Followup.html',context=context)

    # follow up list  
def followup_list(request,id):
    global enquiry
    # enqurid=Enquiry.objects.get(id=id)
    student=Enquiry.objects.filter(id=id).only("Full_Name")
    enquiry=Enquiry.objects.get(id=id)
  
    follow_list=reason_for_follow_up.objects.filter(enquiry_id=id)
    context={'follow_list':follow_list,'enquiry':enquiry,'student':student}

    return render(request,'enquiry/follow_list.html',context)  

def status_chnage(request,id):
   
    enquiry=Enquiry.objects.get(id=id)
   
    # print("Enquri is:",enquiry)
    myreason=Reason_follow.objects.all()
    # follow=reason_for_follow_up.objects.all()
    follow_up=reason_for_follow_up.objects.filter(enquiry_id=id)
    follow_up3=reason_for_follow_up.objects.update(Followup_Status = 'FollowUp Done')
    
    for fo in follow_up:
        print("Folowup status is:",fo.Followup_Status)
    if request.method == 'POST':
            if reason_for_follow_up.objects.count() >= 3:
              if reason_for_follow_up.objects.filter(enquiry_id=id).count()>=3: 
                messages.error(request, 'Sorry, you have reached the maximum number of follow-up submissions.')
                return redirect('enq:follwuplist',id=id)
            enquiry=enquiry
            date = request.POST['date']
            time = request.POST['time']
            remark = request.POST['remark']
            reason=request.POST['reason']
            followup=reason_for_follow_up.objects.create(enquiry=enquiry,date=date,remark=remark,time=time,reason_id=reason)
          
            myreason.reason_id=reason
        
            followup.save()
        
            messages.success(request, 'Thank you for your follow-up submission.')
            return redirect('emanage:follwuplist',id=id)
          
          
    return render(request,'enquiry/Followup1.html',context={'myreason':myreason,'enquiry':enquiry,'follow_up':follow_up})
def enquiry_status(request,id):
    enq_status=Enquiry.objects.get(id=id)
    enq_status1=Enquiry.objects.filter(id=id).update(enquiry_Status = 'Finalized')
    if request.method == "POST":
        name = request.POST.get('Full_Name')
        gender = request.POST.get('gender')
        email = request.POST.get('Email_Address')
        mobile = request.POST.get('Contact_Number')
        address=request.POST.get('Permanent_Address')
        whatsapp = request.POST.get('Whatsapp_Number')
        qualification=request.POST.get('qualification')
        media=request.POST.get('How_did_you_hear_about_us')
        skill = request.POST.get('Technical_Skills')
        certificate = request.POST.get('Certification_Done')
        interest = request.POST.get('Area_of_Intrest')
        reference = request.POST.get('reference')
        remark = request.POST.get('remark')
        Cource_name=request.POST.get('Cource_name')
        address=request.POST.get('Permanent_Address')
        dob=request.POST.get('Date_of_Birth')
        alternate=request.POST.get('Alternate_Number')
        location=request.POST.get('Office_Location_Preference')
        mode=request.POST.get('Training_Mode')
        print("Course name",Cource_name)
        enq_status.Full_Name=name
       
        enq_status.gender=gender
        enq_status.Email_Address=email
        enq_status.Contact_Number=mobile
        enq_status.Whatsapp_Number=whatsapp
        enq_status.qualification=qualification
        enq_status.How_did_you_hear_about_us=media
        enq_status.Technical_Skills=skill
        enq_status.Certification_Done=certificate
        enq_status.Area_of_Intrest=interest
        enq_status.reference=reference
        enq_status.Date_of_Birth=dob
        enq_status.Alternate_Number=alternate
        enq_status.Office_Location_Preference=location
        enq_status.Training_Mode=mode
        enq_status.Permanent_Address=address
        enq_status.remark=remark
       
        enq_status.save()
       
        return redirect('/enq_list')
    else:
        return render(request, 'enquiry/update_enq2.html',context={'enq_status':enq_status})
    
def whatsapp_status(request,id):
    # wapp_status=Enquiry.objects.get(id=id)
    wapp_status1=Enquiry.objects.filter(id=id).update(enquiry_Status = 'Informed_On_whatapp')
    return redirect('/enq_list',id=id)
def excelsheets(request):
    return render(request,'enquiry/excel.html')

def account_page(request,id):
    global fees
    # course=Cource.objects.all()
    enquiry=Enquiry.objects.get(id=id)
    course=Cource.objects.filter(enquiry=id).last()
    if request.method == 'POST':
            enquiry = enquiry
            fname = request.POST.get('Full_Name')
          
            Cource_name=request.POST.get('Cource_name')
            #    fees=request.POST.get('fees')
            total_installment=request.POST['total_installment']
         
        #    pay_date=request.POST.get('pay_date')
            payment_mode=request.POST.get('payment_mode')
            total_amount=request.POST.get('total_amount')
           
            try:
                first_installment=request.POST['installment_1']
            except MultiValueDictKeyError:
                first_installment = None
            try:
               second_installment= request.POST['installment_2']
               second_date=request.POST['date_2']
            except MultiValueDictKeyError:
               second_installment = None
               second_date = None
            
            try:
               third_installment=request.POST['installment_3']
               third_date=request.POST['date_3']
            except MultiValueDictKeyError:
               third_installment = None
               third_date = None

            try:
                four_installment=request.POST['installment_4']
                four_date=request.POST['date_4']
            except MultiValueDictKeyError:
                four_installment = None
                four_date = None
                
            acc=Account.objects.create(enquiry=enquiry,total_installment=total_installment,first_installment=first_installment,payment_mode=payment_mode,second_installment=second_installment,second_date=second_date,third_installment=third_installment,third_date=third_date,four_installment=four_installment,four_date=four_date,total_amount=total_amount)
            enquiry.fname=fname
           
            acc.save()
            messages.success(request,"Form is Submitted Succefully!!")
            return redirect('emanage:home')
    else:
        return render(request,'enquiry/Accounts.html',context={'enquiry':enquiry,'course':course})

def import_data(request):
    if "GET" == request.method:
        return render(request, 'enquiry/excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        active_sheet = wb.active
        print(active_sheet)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'enquiry/enquiry_list.html', {"excel_data":excel_data})
# def import_data(request):
#     if request.method == 'POST' and request.FILES['excel_file']:
#         excel_file = request.FILES['excel_file']
#         workbook = openpyxl.load_workbook(excel_file)
#         sheet = workbook.active

#         for row in sheet.iter_rows(min_row=20, values_only=True):
#             Full_Name = row[1]
#             gender = row[2]
#             Email_Address = row[3]
#             Contact_Number = row[4]
#             Permanent_Address = row[5]
#             Date_of_Birth = row[6]
#             Whatsapp_Number = row[8]
#             Alternate_Number = row[9]
#             How_did_you_hear_about_us = row[10]
#             qualification = row[11]
#             Technical_Skills = row[12]
#             Certification_Done = row[13]
#             Area_of_Intrest = row[14]
#             reference = row[15]
#             remark = row[16]
#             Training_Mode = row[17]
#             Office_Location_Preference = row[18]

#             obj=Enquiry.objects.create(
#                 Full_Name=Full_Name,
#                 gender=gender,
#                 Email_Address=Email_Address,
#                 Contact_Number=Contact_Number,
#                 Permanent_Address=Permanent_Address,
#                 Date_of_Birth=Date_of_Birth,
#                 Whatsapp_Number=Whatsapp_Number,
#                 Alternate_Number=Alternate_Number,
#                 How_did_you_hear_about_us=How_did_you_hear_about_us,
#                 qualification=qualification,
#                 Technical_Skills=Technical_Skills,
#                 Certification_Done=Certification_Done,
#                 Area_of_Intrest=Area_of_Intrest,
#                 reference=reference,
#                 remark=remark,
#                 Training_Mode=Training_Mode,
#                 Office_Location_Preference=Office_Location_Preference
#             )
#             obj.save()
#         return redirect('/enq_list')
    
#     return render(request, 'enquiry/excel.html')
              
     
#############################Admin##################################################
def is_admin(self):
    if str(self.user_type) == 'Is admin':
        return True
    else:
        return False
rec_login_required = user_passes_test(lambda u: True if u.is_admin else False, login_url='/')

def admin_login_required(view_func):
    decorated_view_func = login_required(rec_login_required(view_func), login_url='/')
    return decorated_view_func

@admin_login_required
def admin_page(request):
    all_enquiry=Enquiry.objects.all().count()
    all_followup=reason_for_follow_up.objects.all().count()
    notdone= reason_for_follow_up.objects.filter(Followup_Status = 'pending').count()
    approved = reason_for_follow_up.objects.filter(Followup_Status = 'FollowUp Done').count()
   
    accounts = Account.objects.all()

    # Count the number of occurrences for each installment category
    installment_counts = dict()
    for account in accounts:
        installment = account.payment_status
        installment_counts[installment] = installment_counts.get(installment, 0) + 1

    # Prepare data for the pie chart
    labels = list(installment_counts.keys())
    values = list(installment_counts.values())

    # Create a pie chart using Plotly
    fig = go.Figure(data=go.Pie(labels=labels, values=values))

    # Customize the chart layout
    fig.update_layout(title='Installment Pie Chart')

    # Convert the Plotly chart to HTML
    chart_html = fig.to_html(full_html=False, default_width='100%', default_height='500px')
    context={'all_enquiry':all_enquiry,
            'all_followup':all_followup,
            'notdone':notdone,
            'approved':approved,
            'chart_html':chart_html}
    return render(request,'admin/base.html',context=context)

def enquiry_list(request):
    items_per_page = 9  # Number of items to display per page
    enquiry_list=Enquiry.objects.all()
    total_enquiries=enquiry_list.count()
    paginator = Paginator(enquiry_list, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # Get the
    # enquiry_list=Enquiry.objects.all().order_by('-id').values()
    course=Cource.objects.all()
    total_enquiries=enquiry_list.count()
    return render(request,'admin/enquiry_list.html',{'total_enquiries':total_enquiries,'enquiry_list':page_obj,'course':course,'page_obj':page_obj})

def follo_list(request,id):
    global enquiry
    # enqurid=Enquiry.objects.get(id=id)
    student=Enquiry.objects.filter(id=id).only("Full_Name")
    enquiry=Enquiry.objects.get(id=id)
    follow_list=reason_for_follow_up.objects.filter(enquiry_id=id)
    return render(request,'admin/follow_list.html',context={'enquiry':enquiry,'follow_list':follow_list,'student':student})

  
def follow_pending(request):
    notdone= reason_for_follow_up.objects.filter(Followup_Status = 'pending').all()
    return render(request, 'admin/pend_follo.html', context={'notdone':notdone})

def follow_done(request):
    done= reason_for_follow_up.objects.filter(Followup_Status = 'FollowUp Done').all()
    return render(request, 'admin/done_follo.html', context={'done':done})

def acnt_list(request):
    items_per_page = 9  # Number of items to display per page
    acc=Account.objects.all()
    paginator = Paginator(acc, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  
    return render(request,'admin/account_list.html',context={'acc':page_obj,'page_obj':page_obj})

def fees_list(request,id):
    enquiry=Enquiry.objects.filter(id=id).only("Full_Name")
    obj=Cource.objects.filter(enquiry=id).last()
    fees_record=Account.objects.filter(enquiry_id=id).all()
    # lastrecord=Account.objects.last()
    context={'enquiry':enquiry,'fees_record':fees_record,'obj':obj}
    # return redirect('/feesList',id=id)
    return render(request,'admin/fees_list.html',context)

   
    # fees_record=Account.objects.filter(enquiry_id=id).all()
    # context={'enquiry':enquiry,'fees_record':fees_record,'obj':obj}
    # # return redirect('/feesList',id=id)
    # return render(request,'admin/fees_list.html',context)

def total_follo(request):
    total=reason_for_follow_up.objects.all()
    return render(request,'admin/total_follo.html',context={'total':total})

def system_allocated(request):
    enquiry=Enquiry.objects.all()
    approved = Allocation.objects.filter(system_status = 'allocated').all()
     
    items_per_page = 9  # Number of items to display per page
   
    paginator = Paginator(approved, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # G
    return render(request, 'admin/allocate_list.html', context={'approved':page_obj,'enquiry':enquiry,'page_obj':page_obj})

def not_allocated(request):
    approved = Allocation.objects.filter(system_status = 'not_allocated').all()
    return render(request, 'admin/not_allocat.html', context={'approved':approved})

def total_system(request):
    all_data= Assert_create.objects.all().order_by('-id').values()
    return render(request,'admin/system_list.html',context={'all_data':all_data})

def document_list(request):
    approved = Documetation.objects.all()
    return render(request, 'admin/document_list.html', context={'approved':approved})  

def training_list(request):
    items_per_page = 9  # Number of items to display per page
    all_list=Training.objects.all()
    paginator = Paginator(all_list, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # Get the
    return render(request,'admin/training_list.html',context={'all_list':page_obj,'page_obj':page_obj})

def project_list2(request):
    mydata=Project.objects.all() 
    items_per_page = 9  # Number of items to display per page
   
    paginator = Paginator(mydata, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # G
    return render(request,'admin/project_list.html',context={'mydata':page_obj,'page_obj':page_obj})

def pie_chart_view(request):
    # Retrieve data from the database
    accounts = Account.objects.all()

    # Count the number of occurrences for each installment category
    installment_counts = dict()
    for account in accounts:
        installment = account.payment_status
        installment_counts[installment] = installment_counts.get(installment, 0) + 1

    # Prepare data for the pie chart
    labels = list(installment_counts.keys())
    values = list(installment_counts.values())

    # Create a pie chart using Plotly
    fig = go.Figure(data=go.Pie(labels=labels, values=values))

    # Customize the chart layout
    fig.update_layout(title='Installment Pie Chart')

    # Convert the Plotly chart to HTML
    chart_html = fig.to_html(full_html=False, default_width='100%', default_height='500px')

    return render(request, 'Admin/base.html', {'chart_html': chart_html})

def userlist(request):
    all_user=User.objects.all()
    return render(request,'admin/userlist.html',{'all_user':all_user})

################################Acoount Department########################
def is_account(self):    # login decorator
    if str(self.user_type) == 'Is account':
        return True
    else:
        return False
rec_login_required = user_passes_test(lambda u: True if u.is_account else False, login_url='/')

def account_login_required(view_func):
    decorated_view_func = login_required(rec_login_required(view_func), login_url='/')
    return decorated_view_func

@account_login_required
def account_home(request):
    user=User.objects.filter(is_account=request.user.is_account)
    accounts = Account.objects.all()

    # Count the number of occurrences for each installment category
    installment_counts = dict()
    for account in accounts:
        installment = account.payment_status
        installment_counts[installment] = installment_counts.get(installment, 0) + 1

    # Prepare data for the pie chart
    labels = list(installment_counts.keys())
    values = list(installment_counts.values())

    # Create a pie chart using Plotly
    fig = go.Figure(data=go.Pie(labels=labels, values=values))

    # Customize the chart layout
    fig.update_layout(title='Installment Pie Chart')

    # Convert the Plotly chart to HTML
    chart_html = fig.to_html(full_html=False, default_width='100%', default_height='500px')
    return render(request,'enquiry/account_home.html',{'chart_html': chart_html,'user':user})

# def account_home(request):
#     return render(request,'enquiry/account_home.html')
def account_profile(request):
    user = User.objects.get(id=request.user.id)
    account = Account.objects.filter(user=user).first()
  
    context={
        "user": user,
        "account": account
    }
    return render(request, 'enquiry/account_profile.html', context)


def account_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('enq:account_profile')
    else:
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
      
        try:
            customuser =User.objects.get(id=request.user.id)
            customuser.username = username
            customuser.email = email
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()
            messages.success(request, "Profile Updated Successfully")
            return redirect('emanage:account_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('emanage:account_profile')


@account_login_required
def first_installment(request,id):
    # payment=Account.objects.filter(enquiry_id=id).update(payment_status='first_installment')
    enquiry=Enquiry.objects.get(id=id)
    course=Cource.objects.filter(enquiry=id).last()
    data=Account.objects.filter(enquiry=id).last()
    amt=Account.objects.all()
    payment1=Account.objects.filter(payment_status='Pending').last()
    payment1.save()
    if request.method == 'POST':
            enquiry = enquiry
            name = request.POST.get('Full_Name')

            Cource_name=request.POST.get('Cource_name')
            #    fees=request.POST.get('fees')
            total_installment=request.POST['total_installment']
         
            first_installment=request.POST.get('installment_1')
            total_amount=request.POST.get('total_amount')

        #    pay_date=request.POST.get('pay_date')
            payment_mode=request.POST['payment_mode']
            receipt=request.POST['receipt']
            amount=request.POST.get('amount')
            amount_date=request.POST['amount_date']
            account_remark=request.POST['account_remark']
         
            try:
               second_installment= request.POST['installment_2']
               second_date=request.POST['date_2']
            except MultiValueDictKeyError:
               second_installment = None
               second_date = None
            
            try:
               third_installment=request.POST['installment_3']
               third_date=request.POST['date_3']
            except MultiValueDictKeyError:
               third_installment = None
               third_date = None

            try:
                four_installment=request.POST['installment_4']
                four_date=request.POST['date_4']
            except MultiValueDictKeyError:
                four_installment = None
                four_date = None
          
           
            acc=Account.objects.create(enquiry=enquiry,total_installment=total_installment,first_installment=first_installment,payment_mode=payment_mode,receipt=receipt,second_installment=second_installment,second_date=second_date,third_installment=third_installment,third_date=third_date,four_installment=four_installment,four_date=four_date,account_remark=account_remark,amount=amount,amount_date=amount_date,total_amount=total_amount)
            acc.payment_status = 'first_installment'
            acc.amount_remaining = amount
            acc.save()
            enquiry.name=name
            amt.total_amount=total_amount
            messages.success(request,"First Installment is Submitted Succefully!!")
            return redirect('emanage:account_home')
    else:
        return render(request,'enquiry/first_insta.html',context={'enquiry':enquiry,'course':course,'data':data,'payment1':payment1,'amt':amt})
    
def fees_view(request,id):
    enquiry=Enquiry.objects.filter(id=id).only("Full_Name")
    obj=Cource.objects.filter(enquiry=id).last()
    fees_record=Account.objects.filter(enquiry_id=id).all()
    # lastrecord=Account.objects.last()
    context={'enquiry':enquiry,'fees_record':fees_record,'obj':obj}
    return render(request, 'enquiry/fees_view_list.html', context)

def payment_list(request):
    enquiry=Enquiry.objects.all()
    all_account=Account.objects.all().order_by('-id')
    items_per_page = 9  # Number of items to display per page
    # items = Account.objects.all() # Retrieve the items you want to paginate
    paginator = Paginator(all_account, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # Get 
    return render(request,'enquiry/accountlist.html',context={'all_account':page_obj,'page_obj':page_obj,'enquiry':enquiry})

def account_list(request):
    all_account=Account.objects.all()
    return render(request,'enquiry/accountlist.html',context={'all_account':all_account})
  


def second_installment(request,id):
    enquiry=Enquiry.objects.get(id=id)
    data=Account.objects.filter(enquiry=id).last()
    # payment=Account.objects.filter(enquiry_id=id).update(payment_status='second_installment')
    course=Cource.objects.filter(enquiry=id).last()
    data1=Account.objects.filter(enquiry=id).first()
    amt=Account.objects.filter(enquiry=id).last()
    previous_payment = Account.objects.filter(payment_status = 'first_installment').order_by('amount_date').last()
    if previous_payment:
        previous_payment.amount_remaining = amt.total_amount - amt.amount
        previous_payment.save()
    if request.method == 'POST':
            enquiry = enquiry
            name = request.POST.get('Full_Name')
          
            Cource_name=request.POST.get('Cource_name')
            #    fees=request.POST.get('fees')
            total_installment=request.POST['total_installment']
            payment_mode=request.POST['payment_mode']
            receipt=request.POST['receipt']
            amount=request.POST.get('amount')
            amount_date=request.POST['amount_date']
            account_remark=request.POST['account_remark']
            total_amount=request.POST.get('total_amount')
           
            try:
               second_installment= request.POST['installment_2']
               second_date=request.POST['date_2']
            except MultiValueDictKeyError:
               second_installment = None
               second_date = None
            acc=Account.objects.create(enquiry=enquiry,total_installment=total_installment,payment_mode=payment_mode,receipt=receipt,account_remark=account_remark,amount=amount,amount_date=amount_date,second_installment=second_installment,second_date=second_date,total_amount=total_amount)
            enquiry.name=name
          
            amt.amount=amount
            amt.total_amount=total_amount
            acc.payment_status = 'second_installment'
            acc.save()
            messages.success(request,"Second Installment is Submitted Succefully!!")
            return redirect('emanage:feesList',id=id)
    else:
        return render(request,'enquiry/second_insta.html',context={'enquiry':enquiry,'course':course,'data':data,'amt':amt,'data':data,'data1':data1,'previous_payment':previous_payment})
    
def third_installment(request,id):
    enquiry=Enquiry.objects.get(id=id)
    amount=Account.objects.filter(enquiry=id).all()
    data=Account.objects.filter(enquiry=id).last()
    # payment=Account.objects.filter(enquiry_id=id).update(payment_status='third_installment')
    course=Cource.objects.filter(enquiry=id).last()
    data1=Account.objects.filter(enquiry=id).first()
    print('first is:',data1.amount)
    amt1=Account.objects.filter(enquiry=id).all()
    amt=Account.objects.filter(enquiry=id).last()
    print('last ***',amt)
    amount_paid = Account.objects.filter(enquiry=id).aggregate(amount_paid=Sum('amount'))['amount_paid']
    # amount_paid =   amt.amount + data.amount if data else amt.amount
    print("amount paid",amount_paid)
    previous_payment1 = Account.objects.filter(payment_status='second_installment').order_by('amount_date').first()
    # previous_payment1 = Account.objects.filter(payment_status='third_installment').order_by('amount_date').first()

        # If there is a previous payment, subtract the amount paid from the remaining balance
    if previous_payment1:
            previous_payment1.amount_remaining =data.total_amount - amount_paid
            previous_payment1.save()
  
    if request.method == 'POST':
            enquiry = enquiry
            name = request.POST.get('Full_Name')
       
            Cource_name=request.POST.get('Cource_name')
            #    fees=request.POST.get('fees')
            total_installment=request.POST['total_installment']
            total_amount=request.POST.get('total_amount')
         
        #    pay_date=request.POST.get('pay_date')
            payment_mode=request.POST['payment_mode']
            receipt=request.POST['receipt']
            amount=request.POST.get('amount')
            amount_date=request.POST['amount_date']
            # amount_remaining=request.POST['amount_remaining']
            account_remark=request.POST['account_remark']
            third_installment= request.POST.get('installment_3')
            third_date=request.POST.get('date_3')
          
            acc=Account.objects.create(enquiry=enquiry,total_installment=total_installment,payment_mode=payment_mode,receipt=receipt,account_remark=account_remark,amount=amount,amount_date=amount_date,third_installment=third_installment,third_date=third_date,total_amount=total_amount)
            enquiry.name=name
          
            amt.amount=amount
            amt.total_amount=total_amount
            acc.payment_status = 'third_installment'
            acc.save()
          
            messages.success(request,"Third  Installment is Submitted Succefully!!")
            return redirect('emanage:feesList',id=id)
    else:
        return render(request,'enquiry/third_insta.html',context={'enquiry':enquiry,'course':course,'data':data,'amt':amt,'data':data,'data1':data1,'amount_paid':amount_paid,'previous_payment1':previous_payment1})



def four_installment(request,id):
    enquiry=Enquiry.objects.get(id=id)
    data=Account.objects.filter(enquiry=id).last()
    course=Cource.objects.filter(enquiry=id).last()
    # payment=Account.objects.filter(enquiry_id=id).update(payment_status='fourth_installment')
    data1=Account.objects.filter(enquiry=id).first()
    amt=Account.objects.filter(enquiry=id).last()
    print("last amount",amt.amount)
    amount_paid = Account.objects.filter(enquiry=id).aggregate(amount_paid=Sum('amount'))['amount_paid']
    # amount_paid = Account.first_installment + Account.secound_installment + Account.third_installment 
    previous_payment = Account.objects.filter(payment_status ='third_installment').order_by('amount_date').first()
    print("total amt:",amt.total_amount)
    print(" amt paid:",amount_paid)
        # If there is a previous payment, subtract the amount paid from the remaining balance
    if previous_payment:
        previous_payment.amount_remaining = amt.total_amount - amount_paid
        previous_payment.save()
    if previous_payment.amount_remaining == 0:
        previous_payment.payment_status = 'fourth_installment'
        previous_payment.save()
    # if previous_payment.amount_remaining == 0:
    #     previous_payment.payment_status = 'installment_completed'
    #     previous_payment.save()
    if request.method == 'POST':
            enquiry = enquiry
            name = request.POST.get('Full_Name')
         
            Cource_name=request.POST.get('Cource_name')
            #    fees=request.POST.get('fees')
            total_installment=request.POST['total_installment']
            total_amount=request.POST.get('total_amount')
            first_installment=request.POST.get('installment_1')

        #    pay_date=request.POST.get('pay_date')
            payment_mode=request.POST['payment_mode']
            receipt=request.POST['receipt']
            amount=request.POST.get('amount')
            amount_date=request.POST['amount_date']
            amount_remaining=request.POST.get('amount_remaining')
            account_remark=request.POST['account_remark']
            try:
               four_installment= request.POST['installment_4']
               four_date=request.POST['date_4']
            except MultiValueDictKeyError:
               four_installment = None
               four_date = None
               
            acc=Account.objects.create(enquiry=enquiry,total_installment=total_installment,payment_mode=payment_mode,receipt=receipt,account_remark=account_remark,amount=amount,amount_date=amount_date,four_installment=four_installment,four_date=four_date,amount_remaining=amount_remaining,total_amount=total_amount)
            enquiry.name=name
         
            amt.amount=amount
            amt.total_amount
            acc.payment_status = 'fourth_installment'
            acc.save()
        
            messages.success(request,"All Installment are Completed  Succefully!!")
            return redirect('emanage:feesList',id=id)
    else:
        return render(request,'enquiry/four_insta.html',context={'enquiry':enquiry,'course':course,'data':data,'amt':amt,'data':data,'data1':data1,'amount_paid':amount_paid,'previous_payment':previous_payment})

  
def last_installment(request,id):
    # payment=Account.objects.filter(enquiry_id=id).update(payment_status = 'installment_completed')
    enquiry=Enquiry.objects.get(id=id)
    data=Account.objects.filter(enquiry=id).last()
    course=Cource.objects.filter(enquiry=id).last()
    data1=Account.objects.filter(enquiry=id).first()
    amt=Account.objects.filter(enquiry=id).last()
    amount_paid=Account.objects.aggregate(amount_paid=Sum('amount'))['amount_paid']
    previous_payment = Account.objects.filter(payment_status ='fourth_installment').order_by('amount_date').last()

        # If there is a previous payment, subtract the amount paid from the remaining balance
    if previous_payment:
        previous_payment.amount_remaining = amt.total_amount - amount_paid
        previous_payment.save()
    if previous_payment.amount_remaining == 0:
        previous_payment.payment_status = 'installment_completed'
        previous_payment.save()
    if request.method == 'POST':
           
            enquiry = enquiry
            name = request.POST.get('Full_Name')
            Cource_name=request.POST.get('Cource_name')
            #    fees=request.POST.get('fees')
            total_installment=request.POST['total_installment']
            amount=request.POST.get('amount')
            amount_remaining=request.POST.get('amount_remaining')
            total_amount=request.POST.get('total_amount')
            acc=Account.objects.create(enquiry=enquiry,total_installment=total_installment,amount=amount,amount_remaining=amount_remaining,total_amount=total_amount)
            enquiry.name=name
     
            amt.amount=amount
            amt.total_amount
            acc.payment_status = 'installment_completed'
            acc.save()
            messages.success(request,"All Installment are Completed  Succefully!!")
            return redirect('emanage:feesList',id=id)
    else:
        return render(request,'enquiry/last_insta.html',context={'enquiry':enquiry,'course':course,'data':data,'amt':amt,'data':data,'data1':data1,'amount_paid':amount_paid,'previous_payment':previous_payment})

def admissions_chart(request):
    php_no = Enquiry.objects.filter(Cource_name_id=1).count()
    print("total number of php is",php_no)
    py_no=Enquiry.objects.filter(Cource_name_id=2).count()
    ang_no=Enquiry.objects.filter(Cource_name_id=3).count()
    react_no=Enquiry.objects.filter(Cource_name_id=4).count()
    and_no=Enquiry.objects.filter(Cource_name_id=5).count()
    course_list=['PHP','PYTHON','ANGULAR','REACT','ANDROID']
    num_list=[php_no,py_no,ang_no,react_no,and_no]
    context={'course_list':course_list,'num_list':num_list}
    return render(request,'enquiry/admion.html',context)

def final_account(request):
    final = Enquiry.objects.filter(enquiry_Status = 'Finalized').all()
    return render(request,'enquiry/final_account.html',context={'final':final})

def monthly_installment_totals(request):
    enquiry = Enquiry.objects.all()
    selected_month = request.POST.get('selected_month')

    if selected_month:
        selected_month_total = Account.objects.filter(amount_date__month=selected_month).aggregate(total_amount=Sum('amount'))['total_amount']
    else:
        selected_month_total = None

    total_installment = Account.objects.annotate(month=TruncMonth('amount_date')).values('month').annotate(total_amount=Sum('amount')).order_by('month')
    return render(request,'enquiry/monthly_total.html',context={'selected_month':selected_month,'total_amount_for_month':selected_month_total})
################################### IT  Dept ##############################
def is_it(self):    # login decorator
    if str(self.user_type) == 'Is it':
        return True
    else:
        return False
rec_login_required = user_passes_test(lambda u: True if u.is_it else False, login_url='/')

def IThome_login_required(view_func):
    decorated_view_func = login_required(rec_login_required(view_func), login_url='/')
    return decorated_view_func



def it_home(request):
    user=User.objects.filter(is_it=request.user.is_it)
    total=Allocation.objects.all().count()
    allocated = Allocation.objects.filter(system_status = 'allocated').all().count()
    no=Allocation.objects.filter(system_status = 'not_allocated').all().count()
    context={'total':total,'allocated':allocated,'no':no,'user':user}
    return render(request,'IT/it_home.html',context=context)

def IT_profile(request):
    user = User.objects.get(id=request.user.id)
    student = Allocation.objects.filter(user=user).first()
    context={
        "user": user,
        "student": student
    }
    return render(request, 'IT/it_profile.html', context)


def IT_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('emanage:it_profile')
    else:
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
      
        try:
            customuser =User.objects.get(id=request.user.id)
            customuser.username = username
            customuser.email = email
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()

            # student = Enquiry.objects.get(admin=customuser.id)
            # student.address = address
            # student.save()
            
            messages.success(request, "Profile Updated Successfully")
            return redirect('emanage:it_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('emanage:it_profile')

def assert_creation(request):
    cat=Category.objects.all()
    if request.method == 'POST':
        category = request.POST['category']
        a_name = request.POST['a_name']
        a_description = request.POST['a_description']
        make_year = request.POST['make_year']
        model_name = request.POST['model_name']
        serial_no = request.POST['serial_no']
        condition = request.POST['condition']
        system = request.POST['system']
        photos = request.FILES.getlist('asset_photo')
        ass =Desktop.objects.create(
            a_name=a_name, a_description=a_description,category_id=category,
            make_year=make_year, model_name=model_name, serial_no=serial_no,
            condition=condition, system=system
        )
        cat.category_id=category
        # Save each photo to disk and associate it with the Assert_create instance
        for photo in photos:
            photo_path = default_storage.save(photo.name, ContentFile(photo.read()))
            ass.asset_photo = photo_path
        ass.save()
        messages.success(request,"Asset is Created!")
        return redirect('emanage:system_list')
    else:
        return render(request, 'IT/assert_creat.html',context={'cat':cat})

def load_category(request):
    category_id = request.GET.get('category')
    category = Desktop.objects.filter(category_id=category_id)
   
    return render(request, 'IT/dropdown_list.html',context={'category': category})

def load_list(request):
    desktop_id = request.GET.get('desktop')
    system = system_details.objects.filter(desktop_id=desktop_id).all()
    return render(request, 'IT/load.html',context={'system': system})

def final_list(request):
    final = Enquiry.objects.filter(enquiry_Status = 'Finalized').all()
    return render(request,'IT/final_list.html',context={'final':final})

@IThome_login_required
def system_allocation(request,id):
    # desktop = system_details.objects.get(id=id)
    desktop=Desktop.objects.filter(id=id)
    cat=Category.objects.all()
    brand=Desktop.objects.filter(category_id=id).all()
    mysys=system_details.objects.filter(desktop_id=id).all()
    enquiry=Enquiry.objects.get(id=id)
    name=Enquiry.objects.filter(id=id).only("Full_Name")
    ename=Allocation.objects.filter(enquiry_id=id)
    app_status=Allocation.objects.filter(id=id).update(system_status = 'allocated')
    if request.method == 'POST':
       enquiry=enquiry
       category=request.POST['category']
       desktop=request.POST.get('desktop')
       model_name=request.POST['model_name']
       description=request.POST['description']
       serial_no=request.POST['serial_no']
       L_condition=request.POST['L_condition']
       L_purpose=request.POST['L_purpose']
       L_allocation_date=request.POST['L_allocation_date']
       L_photos= request.FILES.getlist('L_photos')
       system=Allocation.objects.create(enquiry=enquiry,model_name=model_name,description=description,serial_no=serial_no,L_condition=L_condition,L_purpose=L_purpose,L_allocation_date=L_allocation_date,category_id=category,desktop_id=desktop)
       for photo in L_photos:
            photo_path = default_storage.save(photo.name, ContentFile(photo.read()))
            system.L_photos = photo_path
            system.save()
       system.save()
       messages.success(request,"System Allocated ")
       return redirect('emanage:it_home')
    else:
      
       return render(request,'IT/system.html',context={'cat':cat,'enquiry':enquiry,'name':name,'brand':brand,'desktop':desktop,'ename':ename,'mysys':mysys,'app_status':app_status})

def Edit_system(request,id):
    desktop=Desktop.objects.filter(id=id)
    cat=Category.objects.all()
    brand=Desktop.objects.filter(category_id=id).all()
    mysys=system_details.objects.filter(desktop_id=id).all()
    enquiry=Enquiry.objects.get(id=id)
    name=Enquiry.objects.filter(id=id).only("Full_Name")
    ename=Allocation.objects.get(id=id)
    
    if request.method == 'POST':
       enquiry=enquiry
       category=request.POST['category']
     
       model_name=request.POST['model_name']
       description=request.POST['description']
       serial_no=request.POST['serial_no']
       L_condition=request.POST['L_condition']
       L_purpose=request.POST['L_purpose']
       L_allocation_date=request.POST['L_allocation_date']
       L_photos= request.FILES.getlist('L_photos')
       ename.enquiry=enquiry
       ename.category_id=category
       ename.model_name=model_name
       ename.description=description
       ename.serial_no=serial_no
       ename.L_condition=L_condition
       ename.L_purpose=L_purpose
       ename.L_allocation_date=L_allocation_date
       if L_photos:
            ename.L_photos = L_photos[0]  
       ename.save()
       
    return render(request,'IT/update_system.html',context={'desktop':desktop,'ename':ename,'cat':cat,'brand':brand,'name':name,'enquiry':enquiry})

def get_data(request):
    desktop_id = request.GET.get('desktop')
    item =system_details.objects.get(id=desktop_id)
    if item:
        data = {
            'field1': item.model_name,
            'field2': item.description,
            'field3':item.serial_no,
        }
        return JsonResponse(data)
    else:
        return JsonResponse({})


def system_list(request):
    all_data= Assert_create.objects.all()
    items_per_page = 9  # Number of items to display per page
  
    paginator = Paginator(all_data, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  
    return render(request,'IT/system_list.html',context={'all_data':page_obj,'page_obj':page_obj})


def allocation_list(request):
    sys_list=Allocation.objects.all()
    items_per_page = 9  # Number of items to display per page
  
    paginator = Paginator(sys_list, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  
    enq_id=Allocation.objects.select_related('enquiry')
    return render(request,'IT/allocat_list.html',context={'sys_list':page_obj,'page_obj':page_obj})

def update_system(request,id):
    app_status=Allocation.objects.filter(enquiry_id=id).update(system_status = 'allocated')
    return redirect('emanage:allocation_list')

def system_allocated_fun(request):
    approved = Allocation.objects.filter(system_status = 'allocated').all()
    items_per_page = 9  # Number of items to display per page
   
    paginator = Paginator(approved, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # G
    return render(request, 'IT/allocated.html', context={'approved':page_obj,'page_obj':page_obj})

def system_not_allocated(request):
    approved = Allocation.objects.filter(system_status = 'not_allocated').all()
    items_per_page = 9  # Number of items to display per page
   
    paginator = Paginator(approved, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # G
    return render(request, 'IT/not_allocate.html', context={'approved':page_obj,'page_obj':page_obj})

####################### HR Dept ##########################################################
def is_hr(self):    # login decorator
    if str(self.user_type) == 'Is hr':
        return True
    else:
        return False
rec_login_required = user_passes_test(lambda u: True if u.is_hr else False, login_url='/')


def HRhome_login_required(view_func):
    decorated_view_func = login_required(rec_login_required(view_func), login_url='/')
    return decorated_view_func

@HRhome_login_required
def document(request,id):
    doc=Compulsory.objects.all()
    doc1=Optional.objects.all()
    document = Documetation.objects.all()
    enquiry=Enquiry.objects.get(id=id)
    # compulsory=Compulsory.objects.get(id=id)
    # optional=Optional.objects.get(id=id)
    if request.method == 'POST':
        # compulsory=request.POST.get('compulsory')
        # optional=request.POST.get('optional')
        compulsory = request.FILES.getlist('compulsory')
        optional = request.FILES.getlist('optional')
        appointment=request.POST['appointment']
        account_link=request.POST['account_link']
        post_link=request.POST['post_link']
        welcome=request.POST['welcome']
        intro=request.POST['intro']
        idl=0
        for d  in doc:
            idl=d
        opl=0
        for o in doc1:
            opl=o
        adoc=Documetation.objects.create(enquiry=enquiry,appointment=appointment,account_link=account_link,post_link=post_link,welcome=welcome,intro=intro,optional_id=3,compulsory_id=3)
        for photo in compulsory:
            photo_path = default_storage.save(photo.name, ContentFile(photo.read()))
            adoc.compulsory = photo_path
            
            adoc.save()
        for photo in optional:
            photo_path = default_storage.save(photo.name, ContentFile(photo.read()))
            adoc.optional = photo_path
            adoc.save()
        adoc.save()
        document.compulsory_id=compulsory
        document.optional_id=optional
        messages.success(request,"Thank you For Submission!!")
        return redirect('/document_list_1')
    else:
        return render(request,'HR/document1.html',context={'doc':doc,'doc1':doc1,'enquiry':enquiry,'document':document})

def index2_view(request):
    return render(request,'enquiry/index2.html')

def hr_profile(request):
    user = User.objects.get(id=request.user.id)
    hr = Documetation.objects.filter(user=user).first()
    context={
        "user": user,
        "hr": hr
    }
    return render(request, 'HR/hr_profile.html', context)


def hr_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('emanage:hr_profile')
    else:
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
      
        try:
            customuser =User.objects.get(id=request.user.id)
            customuser.username = username
            customuser.email = email
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()
            messages.success(request, "Profile Updated Successfully")
            return redirect('emanage:hr_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('emanage:hr_profile')

@HRhome_login_required
def hr_home(request):
    user=User.objects.filter(is_hr=request.user.is_hr)
    approved = Documetation.objects.all().count()
    return render(request, 'HR/hr_home.html', context={'approved':approved,'user':user})  



def document_view(request):
    approved = Documetation.objects.all()
    items_per_page = 9  # Number of items to display per page
    # items = Account.objects.all() # Retrieve the items you want to paginate
   
    paginator = Paginator(approved, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # Get 
    return render(request, 'HR/doc_list.html', context={'approved':page_obj,'page_obj':page_obj})  

def document_list1(request):
    approved = Documetation.objects.all().count()
 
    return render(request, 'HR/hr_home.html', context={'approved':approved})  

def enq_doc(request):
    enquiry_list = Enquiry.objects.filter(enquiry_Status = 'Finalized').all()
    print("*",enquiry_list)
    return render(request,'HR/document.html',context={'enquiry_list':enquiry_list}) 

def update_doclist(request,id):
  
    doc=Compulsory.objects.all()
    doc1=Optional.objects.all()
    document = Documetation.objects.get(id=id)
    enquiry=Enquiry.objects.get(id=id)
    # compulsory=Compulsory.objects.get(id=id)
    # optional=Optional.objects.get(id=id)
    if request.method == 'POST':
        # compulsory=request.POST.get('compulsory')
        # optional=request.POST.get('optional')
        compulsory = request.FILES.getlist('compulsory')
        optional = request.FILES.getlist('optional')
        appointment=request.POST['appointment']
        account_link=request.POST['account_link']
        post_link=request.POST['post_link']
        welcome=request.POST['welcome']
        intro=request.POST['intro']
        idl=0
        for d  in doc:
            idl=d
        opl=0
        for o in doc1:
            opl=o
        adoc=Documetation.objects.create(enquiry=enquiry,appointment=appointment,account_link=account_link,post_link=post_link,welcome=welcome,intro=intro,optional_id=3,compulsory_id=3)
        for photo in compulsory:
            photo_path = default_storage.save(photo.name, ContentFile(photo.read()))
            adoc.compulsory = photo_path
            
            adoc.save()
        for photo in optional:
            photo_path = default_storage.save(photo.name, ContentFile(photo.read()))
            adoc.optional = photo_path
            adoc.save()
        adoc.save()
        document.compulsory_id=compulsory
        document.optional_id=optional
        messages.success(request,"Thank you For Submission!!")
        return redirect('/document_list_1')
    else:
        return render(request,'HR/document1.html',context={'doc':doc,'doc1':doc1,'enquiry':enquiry,'document':document})

############### Training Department #################################################

def is_training(self):    # login decorator
    if str(self.user_type) == 'Is training':
        return True
    else:
        return False
rec_login_required = user_passes_test(lambda u: True if u.is_training else False, login_url='/')

def Training_login_required(view_func):
    decorated_view_func = login_required(rec_login_required(view_func), login_url='/')
    return decorated_view_func

@Training_login_required
def traning_home(request):
    user=User.objects.filter(is_training=request.user.is_training)
    mydata=Training.objects.filter(training_status='Phase_I_Completed').all().count()
    mydata1=Training.objects.filter(training_status='Phase_II_Completed').all().count()
    mydata2=Training.objects.filter(training_status='Phase_III_Completed').all().count() 
    mydata3=Training.objects.filter(training_status='Phase_IV_Completed').all().count() 
    context={'mydata':mydata,'mydata1':mydata1,'mydata2':mydata2,'mydata3':mydata3,'user':user}
    return render(request,'Training/tr_home.html',context=context)

def training_profile(request):
    user = User.objects.get(id=request.user.id)
    training = Training.objects.filter(user=user).first()
    context={
        "user": user,
        "training": training
    }
    return render(request, 'Training/training_profile.html', context)


def training_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('enq:training_profile')
    else:
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
      
        try:
            customuser =User.objects.get(id=request.user.id)
            customuser.username = username
            customuser.email = email
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()

            # student = Enquiry.objects.get(admin=customuser.id)
            # student.address = address
            # student.save()
            
            messages.success(request, "Profile Updated Successfully")
            return redirect('enq:training_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('enq:training_profile')


@Training_login_required
def training(request,id):
    t=Trainer.objects.all()
    course=Cource.objects.filter(enquiry=id).last()
    enquiry=Enquiry.objects.get(id=id)
    tstatus=Training.objects.filter(enquiry_id=id).update(training_status='Phase_I_Completed')
    if request.method == 'POST':
       enquiry=enquiry
       trainer_name=request.POST.get('trainer_name')
     
       train=Training.objects.create(enquiry=enquiry,trainer_name_id=trainer_name)
       t.trainer_id=trainer_name
       train.save()
       return redirect('emanage:t_status',id=id)
    else:
       return render(request,'Training/training.html',context={'enquiry':enquiry,'t':t,'course':course,'tstatus':tstatus})


def traning_list(request):
    items_per_page = 9  # Number of items to display per page
    # items = Account.objects.all() # Retrieve the items you want to paginate
    all_list=Training.objects.all()
    paginator = Paginator(all_list, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # Get 
    return render(request,'Training/training_list.html',context={'all_list':page_obj,'page_obj':page_obj})

def phase_first(request,id):
    enquiry=Enquiry.objects.get(id=id)
    t_name=Training.objects.filter(enquiry=id).first()
    tstatus=Training.objects.filter(enquiry_id=id).update(training_status='Phase_I_Completed')
    if request.method == 'POST':
       enquiry=enquiry
    #    trainer_name=request.POST.get('trainer_name')
       phase_I=request.POST['phase_I']
       date1=request.POST['date1']
     
       train=Training.objects.create(enquiry=enquiry,phase_I=phase_I,date1=date1)
    #    t_name.trainer_id=trainer_name
       train.save()
       return redirect('emanage:t_status',id=id)
    else:
       return render(request,'Training/phase_first.html',context={'enquiry':enquiry,'t_name':t_name,'tstatus':tstatus})
    

def phase_second(request,id):
    enquiry=Enquiry.objects.get(id=id)
    t_name=Training.objects.filter(enquiry=id).first()
    tstatus=Training.objects.filter(enquiry_id=id).update(training_status='Phase_II_Completed')
    if request.method == 'POST':
       enquiry=enquiry
    #    trainer_name=request.POST.get('trainer_name')
       phase_II=request.POST['phase_II']
       date2=request.POST['date2']
     
       train=Training.objects.create(enquiry=enquiry,phase_II=phase_II,date2=date2)
    #    t_name.trainer_id=trainer_name
       train.save()
       return redirect('emanage:t_status',id=id)
    else:
       return render(request,'Training/phase_second.html',context={'enquiry':enquiry,'t_name':t_name,'tstatus':tstatus})
    

def T_status_list(request,id):
    enquiry=Enquiry.objects.filter(id=id).only("Full_Name")
    obj=Cource.objects.filter(enquiry=id).last()
    trainer=Trainer.objects.all()
    mydata=Training.objects.filter(enquiry_id=id).all()
    all_status=Training.objects.all()
    context={'enquiry':enquiry,'obj':obj,'mydata':mydata,'all_status':all_status,'trainer':trainer}
    return render(request,'Training/training_status_list.html',context=context)

def phase_third(request,id):
    enquiry=Enquiry.objects.get(id=id)
    t_name=Training.objects.filter(enquiry=id).first()
    tstatus=Training.objects.filter(enquiry_id=id).update(training_status='Phase_III_Completed')
    if request.method == 'POST':
       enquiry=enquiry
    #    trainer_name=request.POST.get('trainer_name')
       phase_III=request.POST.get('phase_III')
       date3=request.POST['date3']
       train=Training.objects.create(enquiry=enquiry,phase_III=phase_III,date3=date3)
    #    t_name.trainer_id=trainer_name
       train.save()
       return redirect('emanage:t_status',id=id)
    else:
       return render(request,'Training/phase_third.html',context={'enquiry':enquiry,'t_name':t_name,'tstatus':tstatus})
    
def phase_four(request,id):
    enquiry=Enquiry.objects.get(id=id)
    t_name=Training.objects.filter(enquiry=id).first()
    tstatus=Training.objects.filter(enquiry_id=id).update(training_status='Phase_IV_Completed')
    if request.method == 'POST':
       enquiry=enquiry
    #    trainer_name=request.POST.get('trainer_name')
       phase_IV=request.POST.get('phase_IV')
       date4=request.POST['date4']
       train=Training.objects.create(enquiry=enquiry,phase_IV=phase_IV,date4=date4)
    #    t_name.trainer_id=trainer_name
       train.save()
       return redirect('emanage:project_allocation',id=id)
    else:
       return render(request,'Training/phase_four.html',context={'enquiry':enquiry,'t_name':t_name,'tstatus':tstatus})

def total_phase_I(request):
    mydata=Training.objects.filter(training_status='Phase_I_Completed').all() 
    return render(request,'Training/total_phase_first.html',context={'mydata':mydata})

def total_phase_II(request):
    mydata=Training.objects.filter(training_status='Phase_II_Completed').all() 
    return render(request,'Training/total_phase_second.html',context={'mydata':mydata})

def total_phase_III(request):
    mydata=Training.objects.filter(training_status='Phase_III_Completed').all() 
    return render(request,'Training/total_phase_third.html',context={'mydata':mydata})

def total_phase_IV(request):
    mydata=Training.objects.filter(training_status='Phase_IV_Completed').all() 
    return render(request,'Training/total_phase_four.html',context={'mydata':mydata})


def project_allocation(request,id):
    enquiry=Enquiry.objects.get(id=id)
    course=Cource.objects.filter(enquiry=id).last()
    if request.method == 'POST':
        topic_name=request.POST['topic_name']
        description=request.POST['description']
        date=request.POST['date']
        image=request.FILES.getlist('image')
        project=Project.objects.create(enquiry=enquiry,topic_name=topic_name,description=description,date=date)
        for photo in image:
            photo_path = default_storage.save(photo.name, ContentFile(photo.read()))
            project.image = photo_path
            project.save()
        project.save()
        messages.success(request,"Project Is Allocated!!")
        return redirect('emanage:project_list')
    else:
        return render(request,'Training/project_allocation.html',context={'enquiry':enquiry,'course':course})
  
def project_list(request):
    mydata=Project.objects.all() 
    items_per_page = 9  # Number of items to display per page
    # items = Account.objects.all() # Retrieve the items you want to paginate
  
    paginator = Paginator(mydata, items_per_page)  # Create a Paginator object
    page_number = request.GET.get('page')  # Get the current page number from the request parameters
    page_obj = paginator.get_page(page_number)  # 
    return render(request,'Training/project_list1.html',context={'mydata':page_obj,'page_obj':page_obj})

def project_status_list(request,id):
    enquiry=Enquiry.objects.filter(id=id).only("Full_Name")
    obj=Cource.objects.filter(enquiry=id).last()
    trainer=Trainer.objects.all()
    mydata=Project.objects.filter(enquiry_id=id).all()
    all_status=Project.objects.all()
    context={'enquiry':enquiry,'obj':obj,'mydata':mydata,'all_status':all_status,'trainer':trainer}
    return render(request,'Training/project_status_list.html',context=context)

def project_phase1(request,id):
    project=Project.objects.all()
    course=Cource.objects.filter(enquiry=id).last()
    enquiry=Enquiry.objects.get(id=id)
    tstatus=Project.objects.filter(enquiry_id=id).update(project_status='Phase_I_Completed')
    if request.method == 'POST':
       enquiry=enquiry
       phase_I=request.POST['phase_I']
       date1=request.POST['date1']
       train=Project.objects.create(enquiry=enquiry,phase_I=phase_I,date1=date1)
       train.save()
       messages.success(request,"Phase first Completed!!")
       return redirect('emanage:project_status',id=id)
    else:
       return render(request,'Training/project_phase1.html',context={'enquiry':enquiry,'project':project,'course':course,'tstatus':tstatus})
    
def project_phase2(request,id):
    project=Project.objects.all()
    course=Cource.objects.filter(enquiry=id).last()
    enquiry=Enquiry.objects.get(id=id)
    tstatus=Project.objects.filter(enquiry_id=id).update(project_status='Phase_II_Completed')
    if request.method == 'POST':
       enquiry=enquiry
       phase_II=request.POST['phase_II']
       date2=request.POST['date2']
       train=Project.objects.create(enquiry=enquiry,phase_II=phase_II,date2=date2)
       train.save()
       messages.success(request,"Phase Second Completed!!")
       return redirect('emanage:project_status',id=id)
      
    else:
       return render(request,'Training/project_phase2.html',context={'enquiry':enquiry,'project':project,'course':course,'tstatus':tstatus})
    
def project_phase3(request,id):
    project=Project.objects.all()
    course=Cource.objects.filter(enquiry=id).last()
    enquiry=Enquiry.objects.get(id=id)
    tstatus=Project.objects.filter(enquiry_id=id).update(project_status='Phase_III_Completed')
    if request.method == 'POST':
       enquiry=enquiry
       phase_III=request.POST.get('phase_III')
       date3=request.POST['date3']
       train=Project.objects.create(enquiry=enquiry,phase_III=phase_III,date3=date3)
       train.save()
       messages.success(request,"Phase Third Completed!!")
       return redirect('emanage:project_status',id=id)
    else:
       return render(request,'Training/project_phase3.html',context={'enquiry':enquiry,'project':project,'course':course,'tstatus':tstatus})

def project_phase4(request,id):
    project=Project.objects.all()
    course=Cource.objects.filter(enquiry=id).last()
    enquiry=Enquiry.objects.get(id=id)
    tstatus=Project.objects.filter(enquiry_id=id).update(project_status='Phase_IV_Completed')
    if request.method == 'POST':
       enquiry=enquiry
       phase_IV=request.POST.get('phase_IV')
       date4=request.POST['date4']
       train=Project.objects.create(enquiry=enquiry,phase_IV=phase_IV,date4=date4)
       train.save()
       messages.success(request,"Phase Four Completed!!")
       return redirect('emanage:project_status',id=id)
      
    else:
       return render(request,'Training/project_phase4.html',context={'enquiry':enquiry,'project':project,'course':course,'tstatus':tstatus})
   
def trainerlist(request):
    all_list=Trainer.objects.all()
    return render(request,'training/trainerlist.html',context={'all_list':all_list})  

@Training_login_required
def trainer_name_add(request):
    if request.method=='POST':
        trainer_name=request.POST.get('trainer_name')  
        trainee=Trainer.objects.create(trainer_name=trainer_name)
        trainee.save()
        messages.success(request,"Trainer Added Scussfully")
        return redirect('/trainerlist')
    else:
        return render(request,'training/add_trainer.html')
    
def add_trainer(request):
  
    enquiry_list = Enquiry.objects.filter(enquiry_Status = 'Finalized').all()
   
    return render(request,'Training/add.html',context={'enquiry_list':enquiry_list}) 

def add_project(request):
    enquiry_list = Enquiry.objects.filter(enquiry_Status = 'Finalized').all()
   
    return render(request,'Training/add_project.html',context={'enquiry_list':enquiry_list})
 
def update_trainer_list(request,id):
    all_data=Trainer.objects.get(id=id)
    if request.method=='POST':
        trainer_name=request.POST.get('trainer_name')  
        all_data.trainer_name=trainer_name
        all_data.save()
        messages.success(request,"Trainer Updated Scussfully")
        return redirect('/trainerlist')
    else:
       return render(request,'Training/update_trainer.html',context={'all_data':all_data})